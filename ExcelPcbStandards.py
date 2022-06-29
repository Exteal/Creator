from openpyxl import load_workbook

class ExcelPcbStandards:
    """handles pcb standards data stored in an .xlsx file
    
    standards are taken from eurocircuit : #link_placeholder#
    """
    def __init__(self):
        self.CLASSNUMBERROW = 2
        self.CLASSDRILLROW = 3
        self.STARTCOLUMN = 2
        self.ENDCOLUMN = 36
        self.ROWSTART = 4
        self.ROWEND = 9

    def getPCBInfo(self,pathToFile, number,char):
        wb = load_workbook(pathToFile,data_only=True)
        sheet = wb.active
        
        classStart = self.findFirstMatchingClassNumber(number, sheet)
        firstCell = self.findClassDrill(classStart, char, sheet)
        pcb_info = self.getArrayInfo(firstCell, sheet)
        return pcb_info
        
    def findFirstMatchingClassNumber(self, number,sheet):
        for i in range(self.STARTCOLUMN, self.ENDCOLUMN):
            if((sheet.cell(row=self.CLASSNUMBERROW, column=i).value) == number):
                return i
        return Exception("This class number doesn't exists")
        
    def findClassDrill(self, classStart, char,sheet):
        for i in range(classStart, classStart+5):
            if((sheet.cell(row=self.CLASSDRILLROW, column=i).value) == char):
                return i
        return Exception("This class drill doesn't exists")

    def getArrayInfo(self, column,sheet):
        info = []
        for i in range(self.ROWSTART, self.ROWEND):
            info.append(sheet.cell(row=i, column=column).value)
        return info
